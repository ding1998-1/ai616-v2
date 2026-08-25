"""会议会前问题收集与议题生成路由。"""

import csv
import io

from fastapi import APIRouter, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse

from backend.dependencies import require_meeting, require_user
from backend.models import MeetingAgendaRealtimeCheckRequest, MeetingIssueRequest
from backend.services.meeting_support_service import (
    append_issue,
    generate_agenda,
    get_carryover_todos,
    realtime_check,
)


router = APIRouter(prefix="/api/meetings", tags=["meeting-support"])


@router.get("/{meeting_id}/carryover-todos")
async def carryover_todos(request: Request, meeting_id: str):
    """获取该会议的遗留待办列表。"""
    require_user(request)
    todos = get_carryover_todos(meeting_id)
    return {"success": True, "todos": todos, "total": len(todos)}


def _rows_from_upload(filename: str, raw: bytes) -> list[dict]:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return list(csv.DictReader(io.StringIO(raw.decode("utf-8-sig", errors="replace"))))
    if lower.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=503, detail="当前环境未安装 Excel 解析依赖") from exc
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        worksheet = workbook["问题台账"] if "问题台账" in workbook.sheetnames else workbook.active
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            raise HTTPException(status_code=400, detail="Excel 模板缺少表头")
        keys = [str(value or "").strip() for value in header]
        return [{keys[index]: values[index] if index < len(values) else "" for index in range(len(keys))} for values in worksheet.iter_rows(min_row=2, max_row=301, values_only=True)]
    raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .csv 问题台账")


@router.post("/{meeting_id}/issues")
async def create_issue(request: Request, meeting_id: str, body: MeetingIssueRequest):
    user, _, _ = require_meeting(request, meeting_id)
    try:
        issue, meeting = append_issue(meeting_id, body.model_dump(), user)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    from backend.db import _public_meeting
    return {"success": True, "issue": issue, "meeting": _public_meeting(meeting, include_detail=True)}


@router.get("/issues/template")
async def issue_template(request: Request):
    require_user(request)
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(status_code=503, detail="当前环境未安装 Excel 导出依赖") from exc
    buffer = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "问题台账"
    worksheet.append(["问题描述", "来源部门", "提交人", "发生时间", "关联项目", "涉及金额(万元)", "材料缺口", "备注"])
    worksheet.append(["示例：预算调整方案需要补充测算", "财务部", "张三", "", "", "", "", ""])
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(iter([buffer.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename*=UTF-8''AI会议问题收集模板.xlsx"})


@router.post("/{meeting_id}/issues/import-excel")
async def import_issues(request: Request, meeting_id: str, file: UploadFile = File(...)):
    user, _, _ = require_meeting(request, meeting_id)
    raw = await file.read(20 * 1024 * 1024)
    rows = _rows_from_upload(file.filename or "问题台账.csv", raw)
    imported = []
    for row in rows:
        content = str(row.get("问题描述") or row.get("问题") or row.get("描述") or "").strip()
        if not content:
            continue
        parts = [content]
        for label, key in (("关联项目", "关联项目"), ("涉及金额", "涉及金额(万元)"), ("材料缺口", "材料缺口"), ("备注", "备注")):
            if str(row.get(key) or "").strip():
                parts.append(f"{label}：{str(row[key]).strip()}")
        body = {"name": f"{row.get('来源部门') or ''} {row.get('提交人') or ''}".strip(), "content": "；".join(parts), "source": "excel", "meta": file.filename or ""}
        issue, _ = append_issue(meeting_id, body, user)
        imported.append(issue)
    if not imported:
        raise HTTPException(status_code=400, detail="未读取到有效问题，请检查问题描述列")
    drafts, meeting = generate_agenda(meeting_id, user)
    from backend.db import _public_meeting
    return {"success": True, "importedCount": len(imported), "aiProvider": "local-rule", "issues": imported, "agendaDrafts": drafts, "meeting": _public_meeting(meeting, include_detail=True)}


@router.post("/{meeting_id}/agenda/generate")
async def generate_meeting_agenda(request: Request, meeting_id: str):
    user, _, _ = require_meeting(request, meeting_id)
    try:
        drafts, meeting = generate_agenda(meeting_id, user)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    from backend.db import _public_meeting
    return {"success": True, "aiProvider": "local-rule", "sourceCount": len(meeting.get("issueSources") or []), "agendaDrafts": drafts, "meeting": _public_meeting(meeting, include_detail=True)}


@router.post("/{meeting_id}/agenda/realtime-check")
async def agenda_realtime_check(request: Request, meeting_id: str, body: MeetingAgendaRealtimeCheckRequest):
    user, _, _ = require_meeting(request, meeting_id)
    try:
        results = realtime_check(meeting_id, body.agendaDrafts or [], body.latestTranscripts or [], user)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"success": True, "aiProvider": "local-rule", "meetingId": meeting_id, "results": results}
